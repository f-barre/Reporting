# Librairies
import xlwings as xw
from openpyxl.styles import PatternFill, Color

# Modules / Dépendances
from configuration import APP_CONFIG
# Tools
from tools.date_info import get_year, get_month, get_nb_jours_ouvrables
from tools.find_path import get_agency_path
from tools.format_cell_tools import update_euro, update_float, update_pourcentage, display_border
from tools.safe_actions import dprint, safe_dict_get, get_tableau_size, get_ligne_values, taux


def fill_business_table(manager, sheet, ligne):
    """
    Remplissage du 1er tableau du reporting
    :param manager:
    :param sheet:
    :param ligne:
    :return:
    """
    sheet[f"C{ligne}"].value = manager["manager"]["nom"]
    sheet[f"D{ligne}"] = f"{get_month()}/{get_year()}"


def get_variable(consultant, prime_sheet):
    """
    Calcul de variable d'un consultant
    :param consultant:
    :param prime_sheet:
    :return:
    """
    variable = float(0)
    # Variable = primes + % du CA selon le profil
    # Prime
    # On récupère les dimensions du tableau d'exception
    height = get_tableau_size(prime_sheet, "ligne")
    width = get_tableau_size(prime_sheet, "col")

    # On cherche dans la feuille de primes si le consultant a une somme associée
    for ligne in range(height[0] + 1, height[1] - 1):
        # On récupère les données de la ligne du tableau des primes
        ligne_data = get_ligne_values(prime_sheet, ligne, 2, width[0], width[1])

        # Si les noms corresponds on ajoute les primes entre elles au variable
        if f"{consultant['nom']} {consultant['prenom']}" == f"{ligne_data['Nom']} {ligne_data['Prénom']}":

            for prime in list(ligne_data.items())[2:]:
                # On vérifie que la prime est bien à prendre en compte
                if prime[0] in APP_CONFIG.LISTE_PRIMES_A_COMPTER and prime[1] is not None:
                    variable += prime[1]

    # % du CA
    # Variable: Sénior / manager = 2% du CA total de prod | Directeur = 3% du CA total de prod ==> tout ça *2
    if (consultant.get("fonction", "")).lower() in ["sénior", "senior", "consultant senior", "consultant sénior",
                                                    "manager", "consultant manager"]:
        variable = 0.02 * safe_dict_get(consultant, ["ca total de prod"], fail_result=float(0)) * 2
    elif (consultant.get("fonction", "")).lower() == "directeur":
        variable = 0.03 * safe_dict_get(consultant, ["ca total de prod"], fail_result=float(0)) * 2

    return variable


def fill_consultant_table(manager, sheet, ligne, exception_sheet, prime_sheet):
    """
    Remplissage du tableau des consultants du manager en prenant en compte ses primes ou non
    :param manager:
    :param sheet:
    :param ligne:
    :param exception_sheet:
    :param prime_sheet:
    :return:
    """

    def get_ca_total_de_prod(consultant, exception_sheet):
        """
        Calcul le ca total de prod d'un consultant en prenant en compte s'il est payé au forfait ou non
        :param consultant:
        :param exception_sheet:
        :return:
        """
        ca_total_de_prod = safe_dict_get(consultant, ["ca total de prod"])
        is_exception = False

        # CA total de prod = JTM*nb de jours consommées (valeur contenue dans consultant["ca total de prod"]) mais
        # pour les forfaits il faut prendre le coût lissé (coût du contrat / 12)
        # On récupère les dimensions du tableau d'exception
        height = get_tableau_size(exception_sheet, "ligne")

        # On cherche dans la feuille d'exception si le consultant est renseigné comme exception
        for ligne in range(height[0] + 1, height[1] + 1):
            if exception_sheet[f"B{ligne}"].value == consultant["nom"] and \
                    exception_sheet[f"C{ligne}"].value == consultant["prenom"]:
                ca_total_de_prod = exception_sheet[f"D{ligne}"].value
                is_exception = True

        return {"ca_total_de_prod": ca_total_de_prod, "is_exception": is_exception}

    # On itère sur tous les consultants dirigés par le manager
    for index, consultant in enumerate(manager["consultants"]):
        sheet[f"B{ligne}"].value = f'{safe_dict_get(consultant, ["nom"])} {safe_dict_get(consultant, ["prenom"])}'
        sheet[f"C{ligne}"].value = safe_dict_get(consultant, ["statut"])
        sheet[f"D{ligne}"].value = safe_dict_get(consultant, ["fonction"])
        update_float(sheet[f"E{ligne}"], safe_dict_get(consultant, ["nb de jours en mission"], fail_result=float(0)))

        ca_total_de_prod_result = get_ca_total_de_prod(consultant, exception_sheet)
        update_euro(sheet[f"F{ligne}"], ca_total_de_prod_result["ca_total_de_prod"])
        if ca_total_de_prod_result["is_exception"]:
            sheet[f"B{ligne}"].fill = PatternFill(patternType='solid', fgColor=Color(rgb='6699CC'))

        # Cout de prod
        update_euro(sheet[f"G{ligne}"], safe_dict_get(consultant, ["cout de prod"]))

        # Variable en prime
        update_euro(sheet[f"H{ligne}"], get_variable(consultant, prime_sheet))

        # Marge de prod
        update_euro(sheet[f"I{ligne}"],
                    safe_dict_get(consultant, ["ca total de prod"], fail_result=float(0)) -
                    safe_dict_get(consultant, ["cout de prod"], fail_result=float(0)) - sheet[f"H{ligne}"].value)

        # Taux
        update_pourcentage(sheet[f"J{ligne}"],
                           consultant.get("%", taux(sheet[f"I{ligne}"].value,
                                                    safe_dict_get(consultant, ["ca total de prod"],
                                                                  fail_result=float(0)))))

        update_euro(sheet[f"K{ligne}"], safe_dict_get(consultant, ["ca facture"]))
        sheet[f"L{ligne}"].value = safe_dict_get(consultant, ["fin"])

        # Coloration des cases CA Facturé
        # CA total de prod  > CA Facturé -> rouge
        # CA total de prod  < CA Facturé -> vert
        if sheet[f"K{ligne}"].value < sheet[f"F{ligne}"].value:  # en rouge
            sheet[f"K{ligne}"].fill = PatternFill(patternType='solid', fgColor=Color(rgb='CC0000'))
        elif sheet[f"K{ligne}"].value > sheet[f"F{ligne}"].value:  # en vert
            sheet[f"K{ligne}"].fill = PatternFill(patternType='solid', fgColor=Color(rgb='669900'))

        # Evite d'avoir une ligne vide
        if index is not len(manager["consultants"]) - 1:
            sheet.insert_rows(ligne)

    # Totaux
    nb_consultants = len(manager['consultants'])

    # Supprime la ligne vide par défaut s'il n'y a pas de consultants
    if nb_consultants == 0:
        sheet.delete_rows(ligne)

    # Nb consultant -> Somme
    sheet[f"B{nb_consultants + ligne}"].value = f"=ROWS(B{nb_consultants + ligne - 1}:B{ligne})"

    # Merge 2 cases suivantes pour total
    sheet.merge_cells(f"C{nb_consultants + ligne}:D{nb_consultants + ligne}")

    # Temps consommés -> Moyenne
    sheet[f"E{nb_consultants + ligne}"].value = f"=AVERAGE(E{nb_consultants + ligne - 1}:E{ligne})"

    # CA total de production -> Somme
    sheet[f"F{nb_consultants + ligne}"].value = f"=SUM(F{nb_consultants + ligne - 1}:F{ligne})"

    # Coût de production -> Somme
    sheet[f"G{nb_consultants + ligne}"].value = f"=SUM(G{nb_consultants + ligne - 1}:G{ligne})"

    # Variable consultant -> Somme
    sheet[f"H{nb_consultants + ligne}"].value = f"=SUM(H{nb_consultants + ligne - 1}:H{ligne})"

    # Marge prod -> Somme
    sheet[f"I{nb_consultants + ligne}"].value = f"=SUM(I{nb_consultants + ligne - 1}:I{ligne})"

    # % -> Marge de prod / CA total de prod
    sheet[f"J{nb_consultants + ligne}"].value = f"=I{nb_consultants + ligne}/F{nb_consultants + ligne}"

    # Ca Facturé -> Somme
    sheet[f"K{nb_consultants + ligne}"].value = f"=SUM(K{nb_consultants + ligne - 1}:K{ligne})"


def fill_internes_table(internes, sheet, ligne, end_ligne, prime_sheet):
    """
    Remplissage du tableau des internes du reporting
    :param internes:
    :param sheet:
    :param ligne:
    :param end_ligne:
    :param prime_sheet:
    :return:
    """

    def get_montant_mensuel_charge(interne, prime_sheet):
        """
        Calcul le Montant Mensuel Chargé d'un interne en prenant en compte les jours fériés et le variable
        :param interne:
        :param prime_sheet:
        :return:
        """
        # On récupère le varaible de l'interne (primes etc)
        variable = get_variable(interne, prime_sheet)
        # On récupère le nombre de jours ouvrés du mois
        nb_jours_ouvres_du_mois = get_nb_jours_ouvrables(get_year(), get_month())
        montant_mensuel_charge = ((interne.get("salaire", 0.0) * 1.5) / nb_jours_ouvres_du_mois) * interne.get(
            "nb de jours en interne", 0) + (variable * 1.5)
        return montant_mensuel_charge

    somme_montant_mensuel_charge = 0
    for index, interne in enumerate(internes):
        # Nom interne
        sheet[f"B{ligne + index}"].value = f'{interne.get("nom", "")} {interne.get("prenom", "")}'

        # Montant mensuel chargé [(salaire brut *1.5) / nb de jours ouvrés (21)] * nb de j en interne + (variable*1.5)
        update_euro(sheet[f"C{ligne + index}"], get_montant_mensuel_charge(interne, prime_sheet))

        # Total
        somme_montant_mensuel_charge += interne.get("cout de prod", 0.0)

    sheet[f"C{end_ligne}"].value = somme_montant_mensuel_charge


def fill_equipe_table(equipes, sheet, ligne, end_ligne):
    # TODO: clés fausses car pas encore identifié les équipes
    somme_variable_equipe = 0
    for index, interne in enumerate(equipes):
        # Nom equipe
        sheet[f"H{ligne + index}"].value = interne.get("NOM", "")
        # Marge BU
        sheet[f"I{ligne + index}"].value = interne.get("MARGE BU", "")
        # Salaire fixe
        sheet[f"J{ligne + index}"].value = interne.get("SALAIRE FIXE", "")
        # Variable
        sheet[f"K{ligne + index}"].value = interne.get("VARIABLE", "")
        # Variable equipe
        sheet[f"L{ligne + index}"].value = interne.get("VARIABLE EQUIPE", "")

        # Total
        somme_variable_equipe += interne.get("VARIABLE EQUIPE", "")

    # Merge case Total
    sheet.merge_cells(f"H{end_ligne}:K{end_ligne}")
    sheet[f"L{end_ligne}"].value = somme_variable_equipe


def fill_kpis_mensuels_table(sheet, ligne, lignes_internes_equipes):
    """
    Remplissage du tableau KPIS MENSUELS du reporting
    :param sheet:
    :param ligne:
    :param lignes_internes_equipes:
    :return:
    """
    # Montant
    # CA salarié
    l = ligne - 7 - lignes_internes_equipes
    ca_salarie = 0
    while l >= 4:
        if str(sheet[f"C{l}"].value).lower() == "salarié":
            ca_salarie += sheet[f"F{l}"].value
        l -= 1

    sheet[f"C{ligne}"].value = ca_salarie

    # CA indépendant
    ligne_totaux_consultants = ligne - 6 - lignes_internes_equipes
    sheet[f"C{ligne + 1}"].value = f"=F{ligne_totaux_consultants} - C{ligne}"
    # CA total
    sheet[f"C{ligne + 2}"].value = f"=C{ligne} + C{ligne + 1}"
    # Marge de prod totale
    sheet[f"C{ligne + 3}"].value = f"=I{ligne_totaux_consultants}"
    # Cout interne
    sheet[f"C{ligne + 4}"].value = f"=C{ligne - 3}"
    # Marge BU
    sheet[f"C{ligne + 5}"].value = f"=C{ligne + 3} - C{ligne + 4}"

    # Taux
    # CA salarié
    sheet[f"D{ligne}"].value = f"=C{ligne} / C{ligne + 2}"
    # CA indépendant
    sheet[f"D{ligne + 1}"].value = f"=C{ligne + 1} / C{ligne + 2}"
    # Marge de prod totale
    sheet[f"D{ligne + 3}"].value = f"=J{ligne_totaux_consultants}"
    # Marge BU
    sheet[f"D{ligne + 5}"].value = f"=C{ligne + 5} / C{ligne + 2}"


def fill_kpis_consolides_table(sheet, ligne, agency_name, manager_name):
    """
    Remplissage du tableau KPIS CONSOLIDES du reporting en prenant en
    compte le reporting du mois précédent pour les calcul des cumul (CA par exemple)
    :param sheet:
    :param ligne:
    :param agency_name:
    :param manager_name:
    :return:
    """

    def get_kpis_consolide_ligne(sheet):
        """
        Permet de récupérer la ligne à laquelle démrarre le tableau KPIS CONSOLIDES
        :param sheet:
        :return: ligne de début
        """
        ligne = 8
        while sheet[f"H{ligne}"].value != "CA":
            ligne += 1
        return ligne

    try:
        with xw.App(visible=False) as app:
            # On récupère le dernier reporting et la ligne de son tableau KPIS CONSOLIDES
            previous_reporting_workbook = xw.Book(get_agency_path("previous_reporting", agency_name))
            previous_reporting_sheet = previous_reporting_workbook.sheets[manager_name]
            previous_kpis_consolide_ligne = get_kpis_consolide_ligne(previous_reporting_sheet)

            # CA
            sheet[
                f"I{ligne}"].value = f"=C{ligne + 2} + {previous_reporting_sheet[f'I{previous_kpis_consolide_ligne}'].value}"
            # Marge de production
            sheet[
                f"I{ligne + 1}"].value = f"=C{ligne + 3} + {previous_reporting_sheet[f'I{previous_kpis_consolide_ligne + 1}'].value}"
            # Cout interne
            sheet[
                f"I{ligne + 3}"].value = f"=C{ligne + 4} + {previous_reporting_sheet[f'I{previous_kpis_consolide_ligne + 3}'].value}"
            # Marge BU
            sheet[
                f"I{ligne + 4}"].value = f"=C{ligne + 5} + {previous_reporting_sheet[f'I{previous_kpis_consolide_ligne + 4}'].value}"

    except:
        # CA
        sheet[f"I{ligne}"].value = f"=C{ligne + 2}"
        # Marge de production
        sheet[f"I{ligne + 1}"].value = f"=C{ligne + 3}"
        # Cout interne
        sheet[f"I{ligne + 3}"].value = f"=C{ligne + 4}"
        # Marge BU
        sheet[f"I{ligne + 4}"].value = f"=C{ligne + 5}"

    # Taux marge production
    sheet[f"I{ligne + 2}"].value = f"=I{ligne + 1} / I{ligne}"
    # Taux de marge BU
    sheet[f"I{ligne + 5}"].value = f"=I{ligne + 4} / I{ligne}"


def fill_variable_table(sheet, ligne):
    """
    Remplissage du dernier tableau du reporting, le total de variable
    :param sheet:
    :param ligne:
    :return:
    """
    sheet[f"C{ligne}"].value = f'=C{ligne - 2} / 10'


def creation_managers_sheets(database, workbook, template, exception_sheet, prime_sheet, agency_name):
    """
    Création de toutes les feuilles de manager du reporting
    :param database:
    :param workbook:
    :param template:
    :param exception_sheet:
    :param prime_sheet:
    :param agency_name:
    :return:
    """
    for manager in database:
        dprint(f"Création de la feuille manager de: {manager['manager']['nom']}", priority_level=6)
        dprint(f"Duplication du template de la feuille du manager", priority_level=7)
        new_sheet = workbook.copy_worksheet(template)

        dprint(f"Remplissage du titre de la feuille", priority_level=7)
        new_sheet.title = manager["manager"]["nom"]

        dprint(f"Remplissage du 1er tableau", priority_level=7)
        ligne = 2
        fill_business_table(manager, new_sheet, ligne)
        display_border("B", 2, new_sheet)

        dprint(f"Remplissage du 2ème tableau", priority_level=7)
        ligne += 3  # Saut de ligne inter-tableau / ligne titre / ligne à compléter
        fill_consultant_table(manager, new_sheet, ligne, exception_sheet, prime_sheet)
        display_border("B", 4, new_sheet)

        dprint(f"Remplissage du 3ème / 4ème tableau (côte à côte), internes / équipes", priority_level=7)
        ligne += len(manager["consultants"]) + 3  # ligne totaux / Saut de ligne inter-tableau / ligne titre

        # Four table -> Internes et equipe à la même hauteur => on insert autant de ligne que la hauteur max d'un des 2 tableaux
        max_ligne = max(len(manager["internes"]), len(manager["equipes"]))
        for _ in range(1, max_ligne):
            new_sheet.insert_rows(ligne)

        if (max_ligne == 0):
            new_sheet.delete_rows(ligne)

        fill_internes_table(manager["internes"], new_sheet, ligne, ligne + max_ligne, prime_sheet)
        display_border("B", ligne - 1, new_sheet)
        # fill_equipe_table(manager["equipes"], new_sheet, ligne, ligne + max_ligne) TODO: pas de cas d'équipe, fonction abandonnée
        display_border("H", ligne - 1, new_sheet)

        dprint(f"Remplissage du 5ème / 6ème tableau (côte à côte), KPIs mensuels et KPIs consolidés", priority_level=7)
        ligne += max_ligne + 3
        fill_kpis_mensuels_table(new_sheet, ligne, max_ligne)
        display_border("B", ligne - 1, new_sheet)

        fill_kpis_consolides_table(new_sheet, ligne, agency_name, manager["manager"]["nom"])
        display_border("H", ligne - 1, new_sheet)

        dprint(f"Remplissage du 7ème tableau, variable", priority_level=7)
        ligne += 7
        fill_variable_table(new_sheet, ligne)
        display_border("B", ligne - 1, new_sheet)
