# Librairies
from openpyxl.styles.borders import Border, Side


def format_euro(receive_cell):
    """
    Permet de formater une case au format euro, forme: #,##€
    :param receive_cell:
    :return:
    """
    receive_cell.number_format = '#,##0.00€'


def update_euro(receive_cell, value):
    """
    Permet d'update une case et de la formatter au format euro
    :param receive_cell:
    :param value:
    :return:
    """
    receive_cell.value = value
    format_euro(receive_cell)


def format_pourcentage(receive_cell):
    """
    Permet de formater une case au format %, forme: #,##%
    :param receive_cell:
    :return:
    """
    receive_cell.number_format = '#,##0.00%'


def update_pourcentage(receive_cell, value):
    """
    Permet d'update une case et de la formatter au format %
    :param receive_cell:
    :param value:
    :return:
    """
    receive_cell.value = value
    format_pourcentage(receive_cell)


def format_float(receive_cell, precision=2):
    """
    Permet de formater une case au format float, forme: #,#...#
    Possibilité de choisir la précision de la virgule
    :param receive_cell:
    :return:
    """
    hastags = "".join("#" for k in range(precision))
    zeros = "".join("0" for k in range(precision))
    receive_cell.number_format = f'#,{hastags}0.{zeros}'


def update_float(receive_cell, value, precision=2):
    """
    Permet d'update une case et de la formatter au format float,
    en choisissant la précision
    :param receive_cell:
    :param value:
    :return:
    """
    receive_cell.value = value
    format_float(receive_cell, precision=precision)


def get_table_coords(col_debut_tableau, ligne_start_research, sheet):
    """
    Permet de récupérer les coordonnées d'un tableau à partir de sa case supérieure gauche
    :param col_debut_tableau:
    :param ligne_start_research:
    :param sheet:
    :return: coordonnées
    """
    # Si la colonne est en lettre -> conversion en chiffre
    if isinstance(col_debut_tableau, str):
        col_debut_tableau = ord(col_debut_tableau.lower()) - 96

    """
    cherchons les lignes / colonnes de début fin du tableau
    ************** l_d **************
    *********************************
    * col_debut_tableau ******* c_f *
    *********************************
    ************** l_f **************
    """

    # Ligne de départ du tableau
    l_d = ligne_start_research
    while sheet.cell(row=l_d, column=col_debut_tableau).value is None:
        l_d += 1

    # Ligne de fin du tableau
    l_f = l_d
    while sheet.cell(row=l_f, column=col_debut_tableau).value is not None:
        l_f += 1

    # Colonne de fin du tableau
    c_f = col_debut_tableau
    while sheet.cell(row=l_d, column=c_f).value is not None:
        c_f += 1

    return [col_debut_tableau, l_d, c_f, l_f]


def display_border_table(table_coord, sheet):
    """
    Permet d'afficher les bordures d'un tableau à partir de ses coordonnées
    :param table_coord:
    :param sheet:
    :return:
    """
    for c in range(table_coord[0], table_coord[2]):
        for l in range(table_coord[1], table_coord[3]):
            sheet.cell(row=l, column=c).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )


def display_border(col, ligne, sheet):
    """
    Permet d'afficher les bordures d'un tableau à partir de sa case supérieure gauche
    :param col:
    :param ligne:
    :param sheet:
    :return:
    """
    display_border_table(get_table_coords(col, ligne, sheet), sheet)
